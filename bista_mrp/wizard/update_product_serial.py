import base64
from io import BytesIO
from odoo import models, fields,api
from odoo.exceptions import UserError
import openpyxl

class UpdateProductSerial(models.TransientModel):
    _name = 'update.product.serial'
    _description = 'Sale Add Product Wizard'

    mrp_id = fields.Many2one('mrp.production',string="Default_id")
    operation = fields.Selection([
        ('update_product', 'Product'),
        ('update_serial','To Serial')], string='Operation Type')
    excel_file = fields.Binary(string="Upload Excel File", required=True)
    replace_products_wizard_id = fields.One2many('update.product','wizard_id')
    replace_product_serial_wizard_id = fields.One2many('update.serial','wizard_id')
    
    def read_file(self):
        if self.excel_file:
            if self.operation == 'update_product': 
                self.replace_products_wizard_id.unlink()
                wb = openpyxl.load_workbook(filename=BytesIO(base64.b64decode(self.excel_file)), read_only=True)
                ws = wb.active
                print(ws)
                lines= []
                for record in ws.iter_rows(min_row=2, max_row=None, min_col=None,max_col=None, values_only=True):
                    print(record)
                    mo_no,curr_prod,new_prod = record
                    mo_id = self.env['mrp.production'].search([('name', '=', mo_no)]) 
                    curr_prod_id = self.env['product.product'].search([('default_code', '=', curr_prod)])
                    new_prod_id = self.env['product.product'].search([('default_code', '=', new_prod)])
                    lines.append((0,0,{
                                'mrp_id':mo_id.id,
                                'wizard_id': self.id,
                                'current_product': curr_prod_id.id,
                                'new_product': new_prod_id.id,
                                'state': mo_id.state
                                
                            }))
                self.replace_products_wizard_id = lines 
            else:
                self.replace_products_wizard_id.unlink()
                wb = openpyxl.load_workbook(filename=BytesIO(base64.b64decode(self.excel_file)), read_only=True)
                ws = wb.active
                print(ws)
                lines= []
                for record in ws.iter_rows(min_row=2, max_row=None, min_col=None,max_col=None, values_only=True):
                    print(record)
                    mo_no,curr_prod,old_serial,new_serial = record
                    mo_id = self.env['mrp.production'].search([('name', '=', mo_no)]) 
                    curr_prod_id = self.env['product.product'].search([('default_code', '=', curr_prod)])
                    lines.append((0,0,{
                                'mrp_id':mo_id.id,
                                'wizard_id': self.id,
                                'current_product': curr_prod_id.id,
                                'old_serial': old_serial,
                                'new_serial': new_serial,
                            }))
                self.replace_product_serial_wizard_id = lines 
        else:
            raise UserError("First Please Upload Correct Excel File")
        
        veiw_id = self.env.ref("bista_mrp.view_product_serial_wizard_form").id
        return {
            'name' : "Update Product or Wizard Window",
            'view_mode' : "form",
            'res_model' : 'update.product.serial',
            'res_id' : self.id,
            'view_id' : veiw_id,
            'type' : "ir.actions.act_window",
            'target' : 'new',
            'context': {'default_mrp_id': self.id}
        }
            
    def action_open_wizard(self):
        veiw_id = self.env.ref("bista_mrp.view_product_serial_wizard_form").id
        return {
            'name' : "Update Product or Wizard Window",
            'view_mode' : "form",
            'res_model' : 'update.product.serial',
            'res_id' : self.id,
            'view_id' : veiw_id,
            'type' : "ir.actions.act_window",
            'target' : 'new',
            'context': {'default_mrp_id': self.id}
        }
        
            
    def update_product(self):
        self.read_file()
        for records in self.replace_products_wizard_id:
            for mo_record in records.mrp_id:
                if self.operation == 'update_product':
                    if mo_record.state == 'draft':
                        for product_record in mo_record.move_raw_ids:
                            if product_record.product_id.default_code == records.current_product.default_code:
                                new_product_id = self.env['product.product'].search([('default_code', '=', records.new_product.default_code)]).id
                                product_record.write({'product_id': new_product_id})
                                break
                else:
                    pass
                            
                            
                            
    def update_serial(self):
        self.read_file()
        for records in self.replace_products_wizard_id:
            for mo_record in records.mrp_id:
                if self.operation == 'update_serial':
                    move_id = mo_record.move_raw_ids.filtered(lambda rec : rec.product_id.default_code == records.current_product_code)
                    if move_id:
                        lot_id = move_id.lot_ids.filtered(lambda r : r.name == records.old_serial_number)
                        remove_record = move_id.move_line_ids.filtered(lambda x : x.lot_id in lot_id)
                        new_lot_id = self.env['stock.lot'].search([('name', '=', records.new_serial_number)]).id
                        temp_new_lot_id = self.env['stock.quant'].search([('lot_id', '=', new_lot_id)],limit=1).id
                        if temp_new_lot_id:
                            remove_record.write({'quant_id': temp_new_lot_id})
        
    